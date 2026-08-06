#!/usr/bin/env python3
"""Extract EVERY labelled line of the Commission budget workbook, per country per year.

parse_budget.py takes four totals. This takes the whole hierarchy: every expenditure
programme and every revenue component, for each of the 28 member states, 2000-2024.

Nothing is addressed by fixed index. The header row is the first row carrying ten or
more recognised two-letter country codes; the label is everything to the left of the
first country column; the data are the numeric cells from that column rightwards.

Output is deliberately raw — original Commission label text, no reclassification. The
mapping onto a stable set of funds happens downstream in build_flows.py, where it can
be read and argued with.
"""
import openpyxl, csv, re, os

SRC = "/mnt/user-data/uploads/EU Analysis/eu_budget_spending_and_revenue_2000-2023.xlsx"
OUT = "data/raw/_eu_budget_detail.csv"
CC = {"BE": "BEL", "BG": "BGR", "CZ": "CZE", "DK": "DNK", "DE": "DEU", "EE": "EST",
      "IE": "IRL", "EL": "GRC", "GR": "GRC", "ES": "ESP", "FR": "FRA", "HR": "HRV",
      "IT": "ITA", "CY": "CYP", "LV": "LVA", "LT": "LTU", "LU": "LUX", "HU": "HUN",
      "MT": "MLT", "NL": "NLD", "AT": "AUT", "PL": "POL", "PT": "PRT", "RO": "ROU",
      "SI": "SVN", "SK": "SVK", "FI": "FIN", "SE": "SWE", "UK": "GBR"}

# Which block of the sheet a row belongs to. The sheets run expenditure first, then (from
# 2021) a NextGenerationEU block, then revenue. NGEU is tracked separately because those
# lines are a SUBSET of the expenditure above them, not additional money — double-counting
# them would inflate every 2021+ recipient.
def block_of(seen_total_exp, seen_ngeu, seen_rev, text):
    t = text.upper()
    if "NEXTGENERATIONEU" in t:
        return "ngeu", True
    if t.startswith("REVENUE") or "TOTAL REVENUE" in t:
        return "revenue", True
    return None, False


def _is_top(label, code):
    """Top-level MFF headings. Through 2020 they are the only ALL-CAPS lines; from 2021 the
    workbook lower-cases them but gives them a single-segment code."""
    letters = [c for c in label if c.isalpha()]
    # Six letters minimum: "MFA+" is a programme line, not a heading, and treating it as one
    # re-labels every row beneath it in the 2024 sheet.
    if len(letters) >= 6 and all(c.isupper() for c in letters):
        return True
    return bool(code) and "." not in code and len(code) <= 2


wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
out = []
for sheet in wb.sheetnames:
    if not sheet.strip().isdigit():
        continue
    year = int(sheet)
    rows = list(wb[sheet].iter_rows(values_only=True))

    hdr = None
    for i, r in enumerate(rows[:25]):
        hits = [(j, re.sub(r"[^A-Za-z]", "", str(c)).upper()) for j, c in enumerate(r) if c]
        hits = [(j, c) for j, c in hits if len(c) == 2 and c in CC]
        if len(hits) >= 10:
            hdr = (i, hits)
            break
    if not hdr:
        print(f"  {year}: no header row found — skipped")
        continue
    hrow, cols = hdr
    first = min(j for j, _ in cols)

    block, heading = "expenditure", ""
    for rowno, r in enumerate(rows[hrow + 1:]):
        # Only text cells. Between the label and the first country column sit the
        # workbook's own Total / earmarked / other / non-EU / EU-27 aggregates, which are
        # numeric — joining those into the label produced one "distinct label" per value.
        label_cells = [str(c).strip() for c in r[:first]
                       if c is not None and not isinstance(c, (int, float))]
        text = " ".join(label_cells)
        flat = re.sub(r"\s+", " ", text).strip()
        up = flat.upper()

        # Block transitions. Banner text is not reliable across twenty-five years of layout
        # drift, but the running order is: expenditure, TOTAL EXPENDITURE, then (2021 onward)
        # a NextGenerationEU block closed by TOTAL NGEU, then revenue. So the totals mark the
        # boundaries and the NGEU banner is the only banner that has to be recognised.
        if "NEXTGENERATIONEU" in up:
            block = "ngeu"
            continue
        if not flat:
            continue
        # The TOTAL rows are boundaries, but they belong to the block they close -- emit
        # them there and switch afterwards, or the published total lands in revenue and any
        # reconciliation check against it silently finds nothing to compare.
        switch_after = bool(re.match(r"^TOTAL\s+(EXPENDITURE|NGEU)", up))

        # the leading dotted code, where the sheet carries one ("2.1.11", "1.0.1DAG")
        code = ""
        m = re.match(r"^([0-9OS][0-9A-Za-z.]*)\s+(.*)$", flat)
        if m and (("." in m.group(1)) or len(m.group(1)) <= 2):
            code, flat = m.group(1).rstrip("."), m.group(2).strip()

        # The running top-level heading, needed to disambiguate labels the sheet reuses:
        # "Other" appears under Agriculture and twice under Internal Policies, and "of which
        # ERDF" appears under four different objectives.
        if _is_top(flat, code):
            heading = flat

        vals = [(iso, r[j]) for j, cc in cols
                for iso in [CC[cc]] if j < len(r) and isinstance(r[j], (int, float))]
        if not vals:
            continue
        for iso, v in vals:
            # The sheet position is part of the identity. "of which ERDF" and "Other"
            # each appear several times under different parents; keying on the label alone
            # silently merges them and destroys the hierarchy.
            out.append({"year": year, "iso3": iso, "row": rowno, "block": block,
                        "heading": heading, "code": code, "label": flat,
                        "value": round(float(v), 6)})
        if switch_after:
            block = "revenue"

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=["year", "iso3", "row", "block", "heading", "code", "label", "value"])
    w.writeheader()
    w.writerows(out)
print(f"{OUT}: {len(out)} rows")
import collections
print("by block:", collections.Counter(r["block"] for r in out))
