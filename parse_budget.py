#!/usr/bin/env python3
"""Parse the European Commission's EU spending and revenue workbook into tidy rows.

Source: European Commission, "EU spending and revenue - Data 2000-2024",
eu_budget_spending_and_revenue_2000-2023.xlsx, published 25 September 2025.

The workbook has one sheet per year, 2000-2024, and the layout drifts: the header row
moves, the label column moves, and the total labels gain and lose asterisks. So nothing
is addressed by fixed row or column index. The header row is found by looking for a row
carrying ten or more two-letter country codes; the totals are found by matching label
text anywhere in the left-hand cells.

Three series are extracted per member state per year:

  BUDGET.EXPEND        total EU expenditure allocated to that member state
  BUDGET.CONTRIB       total national contributions paid by that member state
  BUDGET.GNI           gross national income, on the Commission's own definition

The Commission publishes GNI in the same workbook, so shares of national income are
computed against the denominator the contributions were actually assessed on rather
than against a World Bank series built on different conventions.

A deliberate omission: the net balance is NOT computed here. It is derived downstream,
because there is more than one defensible convention and the choice needs to be visible.
"""
import openpyxl, csv, re, os

SRC = "/mnt/user-data/uploads/EU Analysis/eu_budget_spending_and_revenue_2000-2023.xlsx"
OUT = "data/raw/_eu_budget.csv"

# Commission country codes -> ISO3. EL is Greece, UK is the United Kingdom.
CC = {"BE": "BEL", "BG": "BGR", "CZ": "CZE", "DK": "DNK", "DE": "DEU", "EE": "EST",
      "IE": "IRL", "EL": "GRC", "GR": "GRC", "ES": "ESP", "FR": "FRA", "HR": "HRV",
      "IT": "ITA", "CY": "CYP", "LV": "LVA", "LT": "LTU", "LU": "LUX", "HU": "HUN",
      "MT": "MLT", "NL": "NLD", "AT": "AUT", "PL": "POL", "PT": "PRT", "RO": "ROU",
      "SI": "SVN", "SK": "SVK", "FI": "FIN", "SE": "SWE", "UK": "GBR"}
NAME = {}
for r in csv.DictReader(open("data/countries.csv", encoding="utf-8")):
    NAME[r["iso3"]] = r["name"]

# Label wording drifts across the 25 sheets, so each target carries every variant seen.
# ADMIN and CUSTOMS are extracted because the headline totals are not comparable without
# them: administration is allocated to whoever hosts the institutions, and customs duties
# are counted as a national contribution under some conventions and not under the
# Commission's own. Both are the difference between rival "net contributor" figures.
TARGETS = [
    ("BUDGET.EXPEND", r"^TOTAL\s+EXPENDITURE",
     "EU expenditure allocated to the member state"),
    ("BUDGET.CONTRIB", r"^TOTAL\s+national\s+contribution",
     "National contributions to the EU budget (excludes customs duties)"),
    ("BUDGET.GNI", r"^Gross National Income \(GNI\)",
     "Gross national income (European Commission definition)"),
    ("BUDGET.ADMIN", r"^(5\.\s*)?ADMINISTRATION\s*$|^European Public Administration\s*$",
     "EU administrative expenditure allocated to the member state"),
    ("BUDGET.CUSTOMS", r"^Customs duties",
     "Customs duties collected for the EU budget (traditional own resources)"),
]


def sheet_rows(ws):
    return list(ws.iter_rows(values_only=True))


def find_header(rows):
    """The header row is the one carrying at least ten two-letter country codes."""
    for i, r in enumerate(rows):
        # Header cells sometimes carry a footnote marker — 2022 has "LU*" because the
        # Commission revised Luxembourg's Horizon figure. Strip anything that is not a
        # letter before matching, or Luxembourg silently vanishes from that year.
        hits = []
        for j, c in enumerate(r):
            if c is None:
                continue
            code = re.sub(r"[^A-Za-z]", "", str(c)).upper()
            if len(code) == 2 and code in CC:
                hits.append((j, code))
        if len(hits) >= 10:
            return i, hits
    return None, []


def find_row(rows, pattern):
    rx = re.compile(pattern, re.I)
    for i, r in enumerate(rows):
        for c in r[:8]:
            if c and isinstance(c, str) and rx.search(re.sub(r"\s+", " ", c.strip())):
                return i
    return None


out, report = [], []
wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
for sheet in wb.sheetnames:
    if not re.fullmatch(r"\d{4}", sheet):
        continue
    year = int(sheet)
    rows = sheet_rows(wb[sheet])
    hdr, cols = find_header(rows)
    if hdr is None:
        report.append(f"{year}: NO HEADER FOUND")
        continue
    got = {}
    for code, pat, name in TARGETS:
        ri = find_row(rows, pat)
        if ri is None:
            report.append(f"{year}: missing {code}")
            continue
        n = 0
        for j, cc in cols:
            v = rows[ri][j] if j < len(rows[ri]) else None
            if isinstance(v, (int, float)):
                out.append({"iso3": CC[cc], "country": NAME.get(CC[cc], CC[cc]),
                            "indicator_code": code, "indicator_name": name,
                            "unit": "EUR million", "year": year, "value": round(float(v), 4),
                            "source": "European Commission, EU spending and revenue "
                                      "2000-2024 workbook (published 25 Sep 2025)",
                            "retrieved": "2026-08-04"})
                n += 1
        got[code] = n
    report.append(f"{year}: {len(cols)} countries | " +
                  " ".join(f"{k.split('.')[1]}={v}" for k, v in got.items()))

os.makedirs("data/raw", exist_ok=True)
with open(OUT, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=["iso3", "country", "indicator_code", "indicator_name",
                                      "unit", "year", "value", "source", "retrieved"])
    w.writeheader()
    w.writerows(out)

print("\n".join(report))
print(f"\n{len(out)} rows -> {OUT}")
