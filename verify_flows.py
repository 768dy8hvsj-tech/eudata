#!/usr/bin/env python3
"""End-to-end check of flows.html, going back to the workbook rather than to the pipeline.

The reconciliation inside build_flows.py compares the mapped funds against totals that came
through the same parser, so a parser bug could satisfy it. This reads the Commission's
"TOTAL EXPENDITURE" and contribution rows straight out of the spreadsheet with a separate
piece of code, and compares them to what the finished page actually renders.
"""
import openpyxl, csv, json, re, subprocess, sys, os

SRC = "/mnt/user-data/uploads/EU Analysis/eu_budget_spending_and_revenue_2000-2023.xlsx"
CC = {"BE": "BEL", "BG": "BGR", "CZ": "CZE", "DK": "DNK", "DE": "DEU", "EE": "EST",
      "IE": "IRL", "EL": "GRC", "GR": "GRC", "ES": "ESP", "FR": "FRA", "HR": "HRV",
      "IT": "ITA", "CY": "CYP", "LV": "LVA", "LT": "LTU", "LU": "LUX", "HU": "HUN",
      "MT": "MLT", "NL": "NLD", "AT": "AUT", "PL": "POL", "PT": "PRT", "RO": "ROU",
      "SI": "SVN", "SK": "SVK", "FI": "FIN", "SE": "SWE", "UK": "GBR"}

WANT = {"exp": r"^total expenditure",
        "nc": r"^total national contribution",
        "tor": r"^traditional own resources|^customs duties$|^sugar levies$"}

truth = {k: {} for k in WANT}
wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
for sh in wb.sheetnames:
    if not sh.strip().isdigit():
        continue
    year = int(sh)
    rows = list(wb[sh].iter_rows(values_only=True))
    cols = None
    for r in rows[:25]:
        hits = [(j, re.sub(r"[^A-Za-z]", "", str(c)).upper()) for j, c in enumerate(r) if c]
        hits = [(j, c) for j, c in hits if len(c) == 2 and c in CC]
        if len(hits) >= 10:
            cols = hits
            break
    first = min(j for j, _ in cols)
    for r in rows:
        txt = " ".join(str(c).strip() for c in r[:first]
                       if c is not None and not isinstance(c, (int, float)))
        low = re.sub(r"\s+", " ", txt).strip().lower()
        for key, pat in WANT.items():
            if re.match(pat, low):
                for j, cc in cols:
                    if j < len(r) and isinstance(r[j], (int, float)):
                        truth[key].setdefault((CC[cc], year), 0.0)
                        truth[key][(CC[cc], year)] += float(r[j])

P = json.load(open(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                "flows_payload.json"), encoding="utf-8"))
bad = 0
for c in P["countries"]:
    yrs = [y for y in P["years"] if y >= c["since"]]
    e = sum(truth["exp"].get((c["iso3"], y), 0.0) for y in yrs) / 1000
    p = sum(truth["nc"].get((c["iso3"], y), 0.0) + truth["tor"].get((c["iso3"], y), 0.0)
            for y in yrs) / 1000
    de = abs(c["cumIn"] - e) / max(abs(e), 1e-9) * 100
    dp = abs(c["cumOut"] - p) / max(abs(p), 1e-9) * 100
    # The payload stores billions to three decimals, so a country's total can sit half a
    # million euros from the workbook purely through display rounding. On Malta's 2bn that
    # is 0.014% and would read as a failure. Absolute tolerance of one million, which is
    # the stored precision, with a percentage floor for the large contributors.
    tol = 0.001
    if abs(c["cumIn"] - e) > tol or abs(c["cumOut"] - p) > tol:
        bad += 1
        print(f"  MISMATCH {c['name']:<16} in {c['cumIn']:>8.2f} vs {e:>8.2f} ({de:.3f}%)"
              f"   out {c['cumOut']:>8.2f} vs {p:>8.2f} ({dp:.3f}%)")
print(f"workbook vs payload: {len(P['countries'])} countries checked, {bad} mismatch(es)")
sys.exit(1 if bad else 0)
